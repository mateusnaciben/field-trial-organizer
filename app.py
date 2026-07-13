import streamlit as st
import streamlit.components.v1 as components
import fitz
import re
import pandas as pd

from supabase import create_client, Client
from datetime import date, timedelta
from io import BytesIO
from uuid import uuid4

import calendar as py_calendar


st.set_page_config(
    page_title="UGA Field Application Planner",
    page_icon="🌱",
    layout="wide"
)


@st.cache_resource
def get_supabase() -> Client:
    return create_client(
        st.secrets["SUPABASE_URL"],
        st.secrets["SUPABASE_KEY"]
    )


supabase = get_supabase()


APP_CODE_DEFAULTS = {
    "A": 30,
    "B": 45,
    "C": 60,
    "D": 75,
    "E": 90,
    "F": 105,
    "G": 120,
    "H": 135
}


st.markdown("""
<style>
.stApp {
    background: #f7f8fa;
}

.app-top {
    background: transparent;
    border-bottom: 3px solid #BA0C2F;
    padding: 8px 0 14px 0;
    margin-bottom: 18px;
    box-shadow: none;
    border-radius: 0;
}

.app-title {
    font-size: 2rem;
    font-weight: 900;
    color: #111827;
    margin-bottom: 4px;
}

.app-subtitle {
    color: #6b7280;
    font-size: 0.95rem;
}

.product-card {
    background: white;
    border: 1px solid #e5e7eb;
    border-left: 5px solid #BA0C2F;
    border-radius: 14px;
    padding: 14px 16px;
    margin-bottom: 10px;
    box-shadow: 0 3px 10px rgba(0,0,0,0.04);
}

.product-name {
    font-size: 0.95rem;
    font-weight: 800;
    color: #111827;
}

.product-amount {
    font-size: 1.55rem;
    font-weight: 900;
    color: #BA0C2F;
}

.info-line {
    color: #4b5563;
    font-size: 0.95rem;
    margin-bottom: 18px;
}

.stButton > button {
    background: #BA0C2F;
    color: white;
    border-radius: 10px;
    border: none;
    font-weight: 700;
}

.stDownloadButton > button {
    background: #111827;
    color: white;
    border-radius: 10px;
    border: none;
    font-weight: 700;
}
</style>
""", unsafe_allow_html=True)


def init_db():
    return None


def fetch_all_rows(
    table_name,
    columns="*",
    page_size=1000
):
    all_rows = []
    start = 0

    while True:
        response = (
            supabase
            .table(table_name)
            .select(columns)
            .range(
                start,
                start + page_size - 1
            )
            .execute()
        )

        batch = response.data or []

        all_rows.extend(batch)

        if len(batch) < page_size:
            break

        start += page_size

    return all_rows


def reset_database():
    (
        supabase
        .table("application_items")
        .delete()
        .gt("id", 0)
        .execute()
    )


def extract_pdf_text_and_lines(uploaded_file):
    uploaded_file.seek(0)

    pdf = fitz.open(
        stream=uploaded_file.read(),
        filetype="pdf"
    )

    full_text = ""
    visual_lines = []

    for page in pdf:
        full_text += page.get_text("text") + "\n"

        words = page.get_text("words")

        rows = {}

        for word_data in words:
            x0, y0, x1, y1, word = word_data[:5]

            y_key = round(y0 / 3) * 3

            rows.setdefault(
                y_key,
                []
            ).append(
                (x0, word)
            )

        for y in sorted(rows.keys()):
            row_words = [
                word
                for x, word in sorted(
                    rows[y],
                    key=lambda item: item[0]
                )
            ]

            line = " ".join(row_words).strip()

            if line:
                visual_lines.append(line)

    return full_text, visual_lines


def extract_trial_info(text):
    trial_id = ""
    title = ""

    trial_match = re.search(
        r"Trial ID:\s*([A-Z0-9\-]+)",
        text
    )

    if trial_match:
        trial_id = trial_match.group(1).strip()

    title_match = re.search(
        r"Title No\.\s*\d+:\s*(.*)",
        text
    )

    if title_match:
        title = title_match.group(1).strip()

    return trial_id, title


def extract_timing_map(text):
    timing = APP_CODE_DEFAULTS.copy()

    timing_match = re.search(
        r"Timing:\s*(.*?)(?:Trial stake color|GA-|Thimet|No cover|$)",
        text,
        re.DOTALL
    )

    if timing_match:
        timing_text = timing_match.group(1)

        matches = re.findall(
            r"([A-Z])\s*=\s*(\d+)\s*DAP",
            timing_text
        )

        for code, dap in matches:
            timing[code] = int(dap)

    return timing


def normalize_unit(unit):
    unit = unit.strip()

    if unit.lower() == "ml":
        return "mL"

    if unit.lower() == "g":
        return "g"

    return unit


def adjust_weekend_date(app_date):
    if app_date.weekday() == 5:
        return app_date - timedelta(days=1)

    if app_date.weekday() == 6:
        return app_date + timedelta(days=1)

    return app_date


def clean_product_name(prefix):
    prefix = re.sub(
        r"^\d+\s+",
        "",
        prefix
    ).strip()

    product_match = re.match(
        r"(.+?)\s+"
        r"\d+(?:\.\d+)?\s+"
        r"(?:LBA/GAL|GA/L|G/L|%|LB/A|LBS/A)\s+"
        r"[A-Z]+",
        prefix,
        flags=re.IGNORECASE
    )

    if product_match:
        return product_match.group(1).strip()

    return prefix.strip()


def parse_arm_spray_plan(
    visual_lines,
    text,
    location,
    planting_date,
    trial_id,
    title
):
    timing_map = extract_timing_map(text)

    rows = []

    current_treatment = ""

    for line in visual_lines:
        clean = line.strip()

        if "Product quantities required" in clean:
            break

        treatment_match = re.match(
            r"^(\d+)\s+",
            clean
        )

        if treatment_match:
            current_treatment = treatment_match.group(1)

        match = re.search(
            r"(?P<prefix>.+?)\s+"
            r"(?P<apps>[A-Z]{1,10})\s+"
            r"(?:\d+\s+)?"
            r"\d+(?:\.\d+)?\s+GAL/AC\s+"
            r"\d+(?:\.\d+)?\s+L\s+"
            r"(?P<amount>\d+(?:\.\d+)?)\s+"
            r"(?P<unit>mL|ml|g|G)\/mx",
            clean
        )

        if not match:
            match = re.search(
                r"(?P<prefix>.+?)\s+"
                r"(?P<apps>[A-Z]{1,10})\s+"
                r"(?P<amount>\d+(?:\.\d+)?)\s+"
                r"(?P<unit>mL|ml|g|G)\/mx",
                clean
            )

        if not match:
            continue

        prefix = match.group("prefix")

        app_codes = match.group("apps")

        amount = float(
            match.group("amount")
        )

        unit = normalize_unit(
            match.group("unit")
        )

        product = clean_product_name(prefix)

        for app_code in app_codes:
            dap = timing_map.get(app_code)

            if dap is None:
                continue

            raw_app_date = (
                planting_date
                + timedelta(days=dap)
            )

            app_date = adjust_weekend_date(
                raw_app_date
            )

            rows.append({
                "Application Date": app_date,
                "Location": location,
                "DAP": dap,
                "App Code": app_code,
                "Trial": trial_id,
                "Treatment": current_treatment,
                "Title": title,
                "Product": product,
                "Amount": amount,
                "Unit": unit
            })

    return rows


def save_rows(
    rows,
    source_file,
    location,
    planting_date
):
    if not rows:
        raise ValueError(
            "No application rows were provided."
        )

    upload_id = str(uuid4())

    payload = []

    for row in rows:
        payload.append({
            "upload_id": upload_id,
            "trial_id": str(
                row["Trial"]
            ),
            "trial_title": str(
                row.get("Title", "") or ""
            ),
            "location": str(location),
            "planting_date": str(
                planting_date
            ),
            "application_date": str(
                row["Application Date"]
            ),
            "dap": (
                int(row["DAP"])
                if row["DAP"] is not None
                else None
            ),
            "app_code": str(
                row["App Code"]
            ),
            "treatment": str(
                row.get("Treatment", "") or ""
            ),
            "product": str(
                row["Product"]
            ),
            "amount": float(
                row["Amount"]
            ),
            "unit": str(
                row["Unit"]
            ),
            "source_file": str(
                source_file
            )
        })

    (
        supabase
        .table("application_items")
        .insert(payload)
        .execute()
    )


def load_items():
    rows = fetch_all_rows(
        "application_items",
        (
            "application_date,"
            "location,"
            "dap,"
            "app_code,"
            "trial_id,"
            "treatment,"
            "product,"
            "amount,"
            "unit"
        )
    )

    columns = [
        "Date",
        "Location",
        "DAP",
        "App",
        "Trial",
        "Treatment",
        "Product",
        "Amount",
        "Unit"
    ]

    if not rows:
        return pd.DataFrame(
            columns=columns
        )

    df = pd.DataFrame(rows)

    df = df.rename(columns={
        "application_date": "Date",
        "location": "Location",
        "dap": "DAP",
        "app_code": "App",
        "trial_id": "Trial",
        "treatment": "Treatment",
        "product": "Product",
        "amount": "Amount",
        "unit": "Unit"
    })

    return (
        df[columns]
        .sort_values([
            "Date",
            "Location",
            "Trial",
            "Treatment",
            "Product"
        ])
        .reset_index(drop=True)
    )


def load_trials():
    rows = fetch_all_rows(
        "application_items",
        (
            "id,"
            "upload_id,"
            "trial_id,"
            "location,"
            "planting_date,"
            "source_file,"
            "created_at"
        )
    )

    columns = [
        "id",
        "Trial",
        "Location",
        "Planting_Date",
        "Source_File",
        "Created_At"
    ]

    if not rows:
        return pd.DataFrame(
            columns=columns
        )

    df = pd.DataFrame(rows)

    df["upload_id"] = df["upload_id"].fillna(
        "legacy-" + df["id"].astype(str)
    )

    trials_df = (
        df
        .groupby(
            "upload_id",
            as_index=False
        )
        .agg({
            "trial_id": "first",
            "location": "first",
            "planting_date": "first",
            "source_file": "first",
            "created_at": "first"
        })
        .rename(columns={
            "upload_id": "id",
            "trial_id": "Trial",
            "location": "Location",
            "planting_date": "Planting_Date",
            "source_file": "Source_File",
            "created_at": "Created_At"
        })
        .sort_values(
            "Created_At",
            ascending=False
        )
        .reset_index(drop=True)
    )

    return trials_df[columns]


def delete_trial(upload_id):
    if str(upload_id).startswith(
        "legacy-"
    ):
        legacy_id = int(
            str(upload_id).replace(
                "legacy-",
                "",
                1
            )
        )

        (
            supabase
            .table("application_items")
            .delete()
            .eq(
                "id",
                legacy_id
            )
            .execute()
        )

        return

    (
        supabase
        .table("application_items")
        .delete()
        .eq(
            "upload_id",
            str(upload_id)
        )
        .execute()
    )


def calendar_summary(df):
    if df.empty:
        return df

    temp = df.copy()

    temp["Date"] = pd.to_datetime(
        temp["Date"]
    ).dt.date

    temp["DAP_Label"] = (
        temp["DAP"]
        .astype(int)
        .astype(str)
        + " DAP ("
        + temp["App"]
        + ")"
    )

    return (
        temp
        .groupby([
            "Date",
            "DAP_Label",
            "Location"
        ])["Trial"]
        .apply(
            lambda values:
            "; ".join(
                sorted(set(values))
            )
        )
        .reset_index()
        .rename(columns={
            "DAP_Label": "DAP",
            "Trial": "Trials"
        })
        .sort_values([
            "Date",
            "Location"
        ])
    )


def product_totals(df):
    if df.empty:
        return df

    return (
        df
        .groupby([
            "Product",
            "Unit"
        ])["Amount"]
        .sum()
        .reset_index()
        .rename(columns={
            "Amount": "Total Amount"
        })
        .sort_values("Product")
    )


def product_totals_export(df):
    if df.empty:
        return df

    return (
        df
        .groupby([
            "Date",
            "Location",
            "Product",
            "Unit"
        ])["Amount"]
        .sum()
        .reset_index()
        .rename(columns={
            "Amount": "Total Amount"
        })
        .sort_values([
            "Date",
            "Location",
            "Product"
        ])
    )


def create_excel_export(df):
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font,
        PatternFill,
        Border,
        Side,
        Alignment
    )
    from openpyxl.utils import (
        get_column_letter
    )

    def clean_export_product_name(name):
        name = str(name).strip()

        patterns = [
            r"\s+\d+(?:\.\d+)?\s*%AW/W\s+WG\s+\d+(?:\.\d+)?\s*oz\s*wt/a.*$",
            r"\s+\d+(?:\.\d+)?\s*%AW/W\s+WG.*$",
            r"\s+\d+(?:\.\d+)?\s*fl\s*oz/a.*$",
            r"\s+\d+(?:\.\d+)?\s*oz/a.*$",
            r"\s+\d+(?:\.\d+)?\s*oz\s*wt/a.*$",
            r"\s+\d+(?:\.\d+)?\s*lb/a.*$",
            r"\s+\d+(?:\.\d+)?\s*LBA/GAL.*$",
            r"\s+\d+(?:\.\d+)?\s*GA/L.*$",
            r"\s+\d+(?:\.\d+)?\s*G/L.*$",
            r"\s+\d+(?:\.\d+)?\s*%.*$"
        ]

        for pattern in patterns:
            name = re.sub(
                pattern,
                "",
                name,
                flags=re.IGNORECASE
            )

        return name.strip()

    df = df.copy()

    df["Product"] = (
        df["Product"]
        .apply(
            clean_export_product_name
        )
    )

    output = BytesIO()

    wb = Workbook()

    ws = wb.active

    wb.remove(ws)

    dark_blue = "156082"
    light_blue = "CAEDFB"
    green = "92D050"
    dark_green = "38761D"
    white = "FFFFFF"
    black = "000000"

    thin = Side(
        style="thin",
        color="000000"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    def style_cell(
        cell,
        fill=None,
        font_color=black,
        bold=False,
        size=11,
        align="center"
    ):
        if fill:
            cell.fill = PatternFill(
                "solid",
                fgColor=fill
            )

        cell.font = Font(
            bold=bold,
            color=font_color,
            size=size
        )

        cell.alignment = Alignment(
            horizontal=align,
            vertical="center",
            wrap_text=True
        )

        cell.border = border

    cal = calendar_summary(df)

    ws = wb.create_sheet("Calendar")

    ws.merge_cells("B2:E2")

    ws["B2"] = "Calendar"

    style_cell(
        ws["B2"],
        fill=white,
        bold=True,
        size=16
    )

    headers = [
        "Date",
        "DAP",
        "Location / Field",
        "Trials"
    ]

    for col, header in enumerate(
        headers,
        start=2
    ):
        cell = ws.cell(
            row=4,
            column=col,
            value=header
        )

        style_cell(
            cell,
            fill=dark_blue,
            font_color=white,
            bold=True
        )

    for row_number, (_, row) in enumerate(
        cal.iterrows(),
        start=5
    ):
        values = [
            pd.to_datetime(
                row["Date"]
            ).strftime("%-d-%b"),
            row["DAP"],
            row["Location"],
            row["Trials"]
        ]

        for column_number, value in enumerate(
            values,
            start=2
        ):
            cell = ws.cell(
                row=row_number,
                column=column_number,
                value=value
            )

            fill = (
                light_blue
                if row_number % 2 == 0
                else white
            )

            style_cell(
                cell,
                fill=fill
            )

    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 62

    totals_sheet = wb.create_sheet(
        "Product Totals"
    )

    totals = (
        df
        .groupby([
            "Date",
            "Location",
            "Product",
            "Unit"
        ])["Amount"]
        .sum()
        .reset_index()
        .rename(columns={
            "Amount": "Total Amount"
        })
        .sort_values([
            "Date",
            "Location",
            "Product"
        ])
    )

    for col, header in enumerate(
        totals.columns,
        start=1
    ):
        cell = totals_sheet.cell(
            row=1,
            column=col,
            value=header
        )

        style_cell(
            cell,
            fill=dark_blue,
            font_color=white,
            bold=True
        )

    for row_number, (_, row) in enumerate(
        totals.iterrows(),
        start=2
    ):
        for column_number, value in enumerate(
            row,
            start=1
        ):
            cell = totals_sheet.cell(
                row=row_number,
                column=column_number,
                value=value
            )

            style_cell(cell)

    for col in range(
        1,
        len(totals.columns) + 1
    ):
        totals_sheet.column_dimensions[
            get_column_letter(col)
        ].width = 22

    raw = wb.create_sheet("Raw Data")

    for col, header in enumerate(
        df.columns,
        start=1
    ):
        cell = raw.cell(
            row=1,
            column=col,
            value=header
        )

        style_cell(
            cell,
            fill=dark_blue,
            font_color=white,
            bold=True
        )

    for row_number, (_, row) in enumerate(
        df.iterrows(),
        start=2
    ):
        for column_number, value in enumerate(
            row,
            start=1
        ):
            cell = raw.cell(
                row=row_number,
                column=column_number,
                value=value
            )

            style_cell(cell)

    for col in range(
        1,
        len(df.columns) + 1
    ):
        raw.column_dimensions[
            get_column_letter(col)
        ].width = 18

    df["Date"] = pd.to_datetime(
        df["Date"]
    ).dt.date

    for app_date in sorted(
        df["Date"]
        .dropna()
        .unique()
    ):
        day_df = df[
            df["Date"] == app_date
        ].copy()

        sheet_name = (
            pd
            .to_datetime(app_date)
            .strftime("%-d-%b")[:31]
        )

        ws = wb.create_sheet(
            sheet_name
        )

        current_row = 3

        groups = day_df.groupby([
            "Location",
            "DAP",
            "App"
        ])

        for (
            location,
            dap,
            app
        ), group_df in groups:
            ws.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=current_row,
                end_column=10
            )

            title_cell = ws.cell(
                row=current_row,
                column=2,
                value="Application details"
            )

            style_cell(
                title_cell,
                fill=green,
                bold=True,
                size=16
            )

            for col in range(3, 11):
                ws.cell(
                    row=current_row,
                    column=col
                ).fill = PatternFill(
                    "solid",
                    fgColor=green
                )

                ws.cell(
                    row=current_row,
                    column=col
                ).border = border

            current_row += 2

            ws.cell(
                row=current_row,
                column=3,
                value="Location:"
            )

            ws.cell(
                row=current_row,
                column=4,
                value=location
            )

            style_cell(
                ws.cell(
                    row=current_row,
                    column=3
                ),
                bold=True,
                align="left"
            )

            style_cell(
                ws.cell(
                    row=current_row,
                    column=4
                ),
                align="left"
            )

            current_row += 1

            ws.cell(
                row=current_row,
                column=3,
                value="DAP:"
            )

            ws.cell(
                row=current_row,
                column=4,
                value=f"{int(dap)} DAP ({app})"
            )

            style_cell(
                ws.cell(
                    row=current_row,
                    column=3
                ),
                bold=True,
                align="left"
            )

            style_cell(
                ws.cell(
                    row=current_row,
                    column=4
                ),
                align="left"
            )

            current_row += 2

            product_summary = (
                group_df
                .groupby([
                    "Product",
                    "Unit"
                ])["Amount"]
                .sum()
                .reset_index()
                .sort_values("Product")
            )

            ws.cell(
                row=current_row,
                column=3,
                value="Product"
            )

            ws.cell(
                row=current_row,
                column=4,
                value="Total Amount"
            )

            style_cell(
                ws.cell(
                    row=current_row,
                    column=3
                ),
                fill=dark_green,
                font_color=white,
                bold=True
            )

            style_cell(
                ws.cell(
                    row=current_row,
                    column=4
                ),
                fill=dark_green,
                font_color=white,
                bold=True
            )

            product_start_row = (
                current_row + 1
            )

            for i, (_, product_row) in enumerate(
                product_summary.iterrows(),
                start=product_start_row
            ):
                ws.cell(
                    row=i,
                    column=3,
                    value=product_row["Product"]
                )

                ws.cell(
                    row=i,
                    column=4,
                    value=(
                        f'{product_row["Amount"]:g} '
                        f'{product_row["Unit"]}'
                    )
                )

                style_cell(
                    ws.cell(
                        row=i,
                        column=3
                    )
                )

                style_cell(
                    ws.cell(
                        row=i,
                        column=4
                    )
                )

            breakdown = (
                group_df
                .groupby([
                    "Trial",
                    "Product",
                    "Unit"
                ])["Amount"]
                .sum()
                .reset_index()
                .sort_values([
                    "Trial",
                    "Product"
                ])
            )

            card_start_row = current_row - 1

            card_col = 6
            card_row = card_start_row
            card_count = 0

            for trial, trial_df in breakdown.groupby(
                "Trial"
            ):
                card_col = (
                    6
                    if card_count % 2 == 0
                    else 8
                )

                if (
                    card_count > 0
                    and card_count % 2 == 0
                ):
                    card_row += 6

                ws.merge_cells(
                    start_row=card_row,
                    start_column=card_col,
                    end_row=card_row,
                    end_column=card_col + 1
                )

                ws.cell(
                    row=card_row,
                    column=card_col,
                    value=trial
                )

                for c in range(
                    card_col,
                    card_col + 2
                ):
                    style_cell(
                        ws.cell(
                            row=card_row,
                            column=c
                        ),
                        bold=True,
                        align="center"
                    )

                r = card_row + 1

                for _, item in trial_df.iterrows():
                    ws.cell(
                        row=r,
                        column=card_col,
                        value=item["Product"]
                    )

                    ws.cell(
                        row=r,
                        column=card_col + 1,
                        value=(
                            f'{item["Amount"]:g} '
                            f'{item["Unit"]}'
                        )
                    )

                    style_cell(
                        ws.cell(
                            row=r,
                            column=card_col
                        ),
                        align="left"
                    )

                    style_cell(
                        ws.cell(
                            row=r,
                            column=card_col + 1
                        ),
                        align="center"
                    )

                    r += 1

                card_count += 1

            current_row = max(
                product_start_row
                + len(product_summary)
                + 4,
                card_row + 8
            )

        for col in range(1, 13):
            ws.column_dimensions[
                get_column_letter(col)
            ].width = 16

        ws.column_dimensions["B"].width = 4
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["F"].width = 22
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 22
        ws.column_dimensions["I"].width = 16

    wb.save(output)

    output.seek(0)

    return output


def render_visual_calendar(
    calendar_data,
    year,
    month
):
    month_calendar = (
        py_calendar
        .Calendar(firstweekday=6)
        .monthdayscalendar(
            year,
            month
        )
    )

    month_name = pd.to_datetime(
        f"{year}-{month}-01"
    ).strftime("%B %Y")

    html = f"""
    <style>
    .calendar-wrapper {{
        width: 100%;
        padding: 10px 0 20px 0;
        font-family: Arial, sans-serif;
    }}

    .calendar-title {{
        font-size: 28px;
        font-weight: 900;
        color: #111827;
        margin-bottom: 18px;
    }}

    .calendar-grid {{
        display: grid;
        grid-template-columns: repeat(7, 1fr);
        gap: 10px;
    }}

    .calendar-header {{
        font-size: 12px;
        font-weight: 900;
        color: #BA0C2F;
        text-align: center;
        padding: 6px 0;
        text-transform: uppercase;
    }}

    .day-card {{
        min-height: 110px;
        border-radius: 14px;
        border: 1px solid #e5e7eb;
        background: #ffffff;
        padding: 10px;
        box-sizing: border-box;
    }}

    .day-card-active {{
        min-height: 110px;
        border-radius: 14px;
        border: 1.5px solid #BA0C2F;
        background: #fff7f8;
        padding: 10px;
        box-sizing: border-box;
        box-shadow: 0 4px 10px rgba(
            186,
            12,
            47,
            0.10
        );
    }}

    .day-number {{
        font-size: 16px;
        font-weight: 900;
        color: #111827;
        margin-bottom: 8px;
    }}

    .dap-pill {{
        display: inline-block;
        background: #BA0C2F;
        color: white;
        font-size: 11px;
        font-weight: 900;
        padding: 4px 7px;
        border-radius: 999px;
        margin-bottom: 5px;
    }}

    .field-text {{
        color: #111827;
        font-size: 12px;
        font-weight: 800;
        margin-bottom: 4px;
    }}

    .trial-text {{
        color: #374151;
        font-size: 11px;
        line-height: 1.3;
        word-break: break-word;
    }}

    .empty-card {{
        min-height: 110px;
        background: transparent;
    }}
    </style>

    <div class="calendar-wrapper">
        <div class="calendar-title">
            📅 {month_name}
        </div>

        <div class="calendar-grid">
    """

    for day_name in [
        "Sun",
        "Mon",
        "Tue",
        "Wed",
        "Thu",
        "Fri",
        "Sat"
    ]:
        html += (
            "<div class='calendar-header'>"
            f"{day_name}"
            "</div>"
        )

    for week in month_calendar:
        for day_number in week:
            if day_number == 0:
                html += (
                    "<div class='empty-card'></div>"
                )

                continue

            current_date = date(
                year,
                month,
                day_number
            )

            rows = calendar_data[
                calendar_data["Date"]
                == current_date
            ]

            if rows.empty:
                html += f"""
                <div class="day-card">
                    <div class="day-number">
                        {day_number}
                    </div>
                </div>
                """

            else:
                html += f"""
                <div class="day-card-active">
                    <div class="day-number">
                        {day_number}
                    </div>
                """

                for _, row in rows.iterrows():
                    trials = row["Trials"]

                    if len(trials) > 52:
                        trials = (
                            trials[:52]
                            + "..."
                        )

                    html += f"""
                    <div class="dap-pill">
                        {row["DAP"]}
                    </div>

                    <div class="field-text">
                        {row["Location"]}
                    </div>

                    <div class="trial-text">
                        {trials}
                    </div>
                    """

                html += "</div>"

    html += """
        </div>
    </div>
    """

    components.html(
        html,
        height=760,
        scrolling=True
    )


init_db()


st.markdown("""
<div class="app-top">
    <div class="app-title">
        🌱 UGA Field Application Planner
    </div>

    <div class="app-subtitle">
        Plant Pathology · Tifton Campus · Field trial application planning
    </div>
</div>
""", unsafe_allow_html=True)


(
    tab_upload,
    tab_calendar,
    tab_date,
    tab_trials,
    tab_export
) = st.tabs([
    "Upload Trial",
    "Calendar",
    "Application Date View",
    "Saved Trials",
    "Export"
])


with tab_upload:
    st.subheader(
        "Upload ARM Spray/Seeding Plan PDF"
    )

    uploaded_file = st.file_uploader(
        "Upload the ARM Spray/Seeding Plan PDF",
        type=["pdf"]
    )

    if uploaded_file:
        text, visual_lines = (
            extract_pdf_text_and_lines(
                uploaded_file
            )
        )

        (
            extracted_trial_id,
            extracted_title
        ) = extract_trial_info(text)

        st.success(
            "PDF loaded successfully"
        )

        trial_id = st.text_input(
            "Trial ID",
            value=extracted_trial_id
        )

        title = st.text_input(
            "Trial Title",
            value=extracted_title
        )

        col1, col2 = st.columns(2)

        with col1:
            location = st.text_input(
                "Location / Field",
                placeholder="Example: Cotton Field"
            )

        with col2:
            planting_date = st.date_input(
                "Planting Date",
                value=date.today()
            )

        if (
            trial_id
            and location
            and planting_date
        ):
            rows = parse_arm_spray_plan(
                visual_lines,
                text,
                location,
                planting_date,
                trial_id,
                title
            )

            st.subheader(
                "Detected Application Items"
            )

            if rows:
                preview_df = pd.DataFrame(
                    rows
                )

                edited_df = st.data_editor(
                    preview_df,
                    width="stretch",
                    num_rows="dynamic",
                    hide_index=True
                )

                if st.button(
                    "Save Trial",
                    type="primary"
                ):
                    try:
                        save_rows(
                            edited_df.to_dict(
                                "records"
                            ),
                            uploaded_file.name,
                            location,
                            planting_date
                        )

                        st.success(
                            "Trial saved successfully."
                        )

                    except Exception as error:
                        st.error(
                            "The trial could not be saved."
                        )

                        st.exception(error)

            else:
                st.error(
                    "No application items were detected."
                )

                with st.expander(
                    "Show extracted PDF lines for debugging"
                ):
                    for line in visual_lines[:120]:
                        st.text(line)


with tab_calendar:
    st.subheader("Master Calendar")

    try:
        df = load_items()

    except Exception as error:
        st.error(
            "Could not load saved applications."
        )

        st.exception(error)

        df = pd.DataFrame()

    if df.empty:
        st.info(
            "No saved trials yet."
        )

    else:
        df["Date"] = pd.to_datetime(
            df["Date"]
        ).dt.date

        available_months = sorted({
            application_date.strftime("%Y-%m")
            for application_date
            in df["Date"]
            .dropna()
            .unique()
        })

        selected_month = st.selectbox(
            "Select month",
            available_months,
            format_func=lambda value:
            pd.to_datetime(
                value + "-01"
            ).strftime("%B %Y")
        )

        year, month = map(
            int,
            selected_month.split("-")
        )

        converted_dates = pd.to_datetime(
            df["Date"]
        )

        month_dates = df[
            (
                converted_dates.dt.year
                == year
            )
            & (
                converted_dates.dt.month
                == month
            )
        ].copy()

        calendar_data = calendar_summary(
            month_dates
        )

        render_visual_calendar(
            calendar_data,
            year,
            month
        )

        st.divider()

        st.subheader(
            "Calendar Table"
        )

        st.dataframe(
            calendar_data,
            width="stretch",
            hide_index=True
        )


with tab_date:
    try:
        df = load_items()

    except Exception as error:
        st.error(
            "Could not load saved applications."
        )

        st.exception(error)

        df = pd.DataFrame()

    if df.empty:
        st.info(
            "No saved applications yet."
        )

    else:
        dates = sorted(
            df["Date"]
            .dropna()
            .unique()
        )

        selected_date = st.selectbox(
            "Select application date",
            dates
        )

        day_df = (
            df[
                df["Date"]
                == selected_date
            ]
            .sort_values([
                "Location",
                "Trial",
                "Treatment",
                "Product"
            ])
        )

        totals_df = product_totals(
            day_df
        )

        date_dap = day_df[
            "DAP"
        ].iloc[0]

        date_app = day_df[
            "App"
        ].iloc[0]

        date_location = day_df[
            "Location"
        ].iloc[0]

        st.markdown(
            "## Application Sheet"
        )

        st.markdown(
            f"""
            <div class="info-line">
                <b>{selected_date}</b>
                · {date_location}
                · {int(date_dap)} DAP ({date_app})
                · {day_df['Trial'].nunique()} trial(s)
                · {totals_df['Product'].nunique()} product(s)
            </div>
            """,
            unsafe_allow_html=True
        )

        col1, col2 = st.columns(
            [0.34, 0.66],
            gap="large"
        )

        with col1:
            st.markdown(
                "### Products to Separate"
            )

            for _, row in totals_df.iterrows():
                st.markdown(
                    f"""
                    <div class="product-card">
                        <div class="product-name">
                            {row['Product']}
                        </div>

                        <div class="product-amount">
                            {row['Total Amount']:g} {row['Unit']}
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

        with col2:
            st.markdown(
                "### Trial Details"
            )

            display_df = day_df[[
                "Trial",
                "Treatment",
                "Product",
                "Amount",
                "Unit",
                "App",
                "DAP"
            ]].copy()

            st.dataframe(
                display_df,
                width="stretch",
                hide_index=True
            )


with tab_trials:
    st.subheader("Saved Trials")

    try:
        trials_df = load_trials()

    except Exception as error:
        st.error(
            "Could not load saved trials."
        )

        st.exception(error)

        trials_df = pd.DataFrame()

    if trials_df.empty:
        st.info(
            "No saved trials yet."
        )

    else:
        st.dataframe(
            trials_df,
            width="stretch",
            hide_index=True
        )

        st.divider()

        st.subheader(
            "Delete Trial Uploaded by Mistake"
        )

        options = {
            (
                f"{row['Trial']} | "
                f"{row['Location']} | "
                f"{row['Source_File']} | "
                f"ID {row['id']}"
            ): row["id"]
            for _, row
            in trials_df.iterrows()
        }

        selected = st.selectbox(
            "Select trial to delete",
            list(options.keys())
        )

        confirm = st.checkbox(
            "Confirm deletion"
        )

        if st.button(
            "Delete selected trial",
            disabled=not confirm
        ):
            try:
                delete_trial(
                    options[selected]
                )

                st.success(
                    "Trial deleted successfully."
                )

                st.rerun()

            except Exception as error:
                st.error(
                    "The trial could not be deleted."
                )

                st.exception(error)

        st.divider()

        st.subheader(
            "Reset Database"
        )

        reset_confirm = st.checkbox(
            "Delete ALL saved data"
        )

        if st.button(
            "Delete everything",
            disabled=not reset_confirm
        ):
            try:
                reset_database()

                st.success(
                    "All saved data deleted."
                )

                st.rerun()

            except Exception as error:
                st.error(
                    "The database could not be cleared."
                )

                st.exception(error)


with tab_export:
    st.subheader("Export Excel")

    try:
        df = load_items()

    except Exception as error:
        st.error(
            "Could not load data for export."
        )

        st.exception(error)

        df = pd.DataFrame()

    if df.empty:
        st.info(
            "No saved data to export."
        )

    else:
        excel_file = create_excel_export(
            df
        )

        st.download_button(
            label="Download Excel File",
            data=excel_file,
            file_name=(
                "Field_Application_Planner.xlsx"
            ),
            mime=(
                "application/"
                "vnd.openxmlformats-"
                "officedocument."
                "spreadsheetml.sheet"
            ),
            type="primary"
        )

        st.markdown(
            "### Preview: Product Totals"
        )

        st.dataframe(
            product_totals_export(df),
            width="stretch",
            hide_index=True
        )
